import json
from openpyxl import load_workbook


def get_day_or_night(shift):
    if "dayOrNight" in shift:
        return shift["dayOrNight"]
    else:
        return "NO DATA"

def generate_job_label(shift_type, day_or_night):
    if shift_type == "Call" or shift_type == "Vacation":
        return shift_type
    elif shift_type == "UC" or shift_type == "UCW" or shift_type == "UCS":
        if day_or_night == "day":
            return "UC:D-O"
        elif day_or_night == "night":
            return "UC:N-O"
        else:
            return "UC-O"
    else:
        raise NameError("Failed to generate label")

def create_new_text(old_txt, new_txt):
    if old_txt == "" or old_txt is None:
        return new_txt
    elif new_txt in old_txt:
        return old_txt
    else:
        return old_txt + ", " + new_txt

def format_date(raw_date):
    split_date = raw_date.split("-")
    return split_date[1] + "/" + split_date[2] + "/" + split_date[0]

def get_year(raw_date):
    split_date = raw_date.split("/")
    return split_date[2]


# Initialize variables
destination_file = "dest2022.xlsx"
source_filename = "shifts_modified.json"
cols_by_date = {}
rows_by_provider = {}
year_to_migrate = "2022"

# Load source data
with open(source_filename) as src_json_file:
    src_data = json.load(src_json_file)

# Open destination file and select sheet
workbook = load_workbook(filename=destination_file)
sheet = workbook.active
date_row = sheet[2]
provider_col = sheet["A"]
shifts_processed = 0
shifts_migrated = 0

# Populate cols index
for i in range(len(date_row)):
    if date_row[i].value != "None":
        cols_by_date[date_row[i].value] = date_row[i].column

# Populate rows index
for j in range(len(provider_col)):
    if provider_col[j].value != "None":
        rows_by_provider[provider_col[j].value] = provider_col[j].row

# Loop through all shifts and write those for selected year to destination file
for k in src_data["shifts"]:
    shift_type = k["shiftType"]
    day_or_night = get_day_or_night(k)
    label = generate_job_label(shift_type, day_or_night)
    date = format_date(k["start"]["$date"].split("T")[0])
    year = get_year(date)
    try:
        provider = k["employee"]["$oid"].split(" ")[1]
    except:
        print("Failed on:")
        print(k["employee"]["$oid"])
    if year == year_to_migrate and label != "Vacation" and label is not None and provider != "McDonald":
        col = cols_by_date[date]
        row = rows_by_provider[provider]
        dest_cell = sheet.cell(row=row, column=col)
        old_value = dest_cell.value
        new_value = create_new_text(old_value, label)
        dest_cell.value = new_value
        shifts_migrated += 1
    shifts_processed += 1

# Save the file
workbook.save(filename=destination_file)
print("Shifts processed: " + str(shifts_processed) + "\nShifts migrated: " + str(shifts_migrated) + "\nMigration Complete")

